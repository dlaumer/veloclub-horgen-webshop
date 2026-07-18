#!/usr/bin/env python3
"""
Import the legacy "Velo Club Horgen Webshop.xlsx" (Lagerbestand + Bestellungen
sheets) into the new PocketBase backend.

Run this on YOUR OWN computer, not the server - it only needs network access
to your public PocketBase URL and the local xlsx file. Nothing here needs
SSH or server access.

Setup (once):
    pip install openpyxl requests

Usage:
    python import_legacy_xlsx.py "C:\\path\\to\\Velo Club Horgen Webshop (3).xlsx"

You'll be prompted for your PocketBase superuser email/password - they are
only used to log in and are never stored or sent anywhere else.

Safe to re-run: articles are matched by article_number and orders by
order_number, so anything already imported is skipped instead of duplicated.
"""

import re
import sys
import openpyxl
import requests

POCKETBASE_URL = "https://api-webshop-veloclubhorgen.duckdns.org"

# Lagerbestand sheet: real header row is 5 (rows 1-4 are a title/banner area)
LAGERBESTAND_HEADER_ROW = 5

# Size columns as they appear in the Lagerbestand header. A cell value of
# "x" (or blank) means "this product doesn't come in that size" and is
# skipped; a number (incl. 0) means "comes in this size, this much stock".
SIZE_COLUMNS = [
    "OneSize", "6/7 Jahre", "8/9 Jahre", "10/11 Jahre", "12/13 Jahre",
    "36-41", "42-47", "XXS", "XS", "S", "M", "L", "XL", "XXL", "3XL",
]


def to_id_str(value):
    """Excel gives us numeric IDs (article numbers, order IDs) as floats."""
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value))
    return str(value).strip()


def parse_image_urls(raw):
    """'[url1, url2, url3]' -> ['url1', 'url2', 'url3']"""
    if not raw:
        return []
    s = str(raw).strip()
    if s.startswith("[") and s.endswith("]"):
        s = s[1:-1]
    return [u.strip() for u in s.split(",") if u.strip()]


def header_map(ws, header_row):
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    return {h: i + 1 for i, h in enumerate(headers) if h}


class PocketBase:
    def __init__(self, base_url):
        self.base_url = base_url.rstrip("/")
        self.token = None

    def login(self):
        # Sanity-check the server is even reachable before asking for
        # credentials, so a network problem fails fast with a clear message
        # instead of hanging silently later.
        print(f"Checking {self.base_url} is reachable...")
        sys.stdout.flush()
        try:
            r = requests.get(f"{self.base_url}/api/health", timeout=10)
            r.raise_for_status()
            print("Server is reachable.")
        except requests.exceptions.RequestException as err:
            print(f"\nCould NOT reach {self.base_url}: {err}")
            print("Check that PocketBase is running and this URL is correct/public")
            print("before continuing. Aborting.")
            sys.exit(1)

        email = input(
            "PocketBase superuser email (leave blank to skip login, e.g. if you "
            "temporarily opened up the collection API rules): "
        ).strip()
        if not email:
            print("Skipping login - requests will be sent unauthenticated.")
            self.token = None
            return

        # Note: using a plain (visible) prompt instead of getpass.getpass()
        # here on purpose - getpass is known to hang indefinitely in Git
        # Bash / MSYS2 terminals on Windows, since it can't read hidden
        # input from that kind of console. This is a local one-off admin
        # script, so a visible password entry is an acceptable trade-off.
        password = input("PocketBase superuser password (visible while typing): ").strip()
        print("Logging in...")
        sys.stdout.flush()
        try:
            r = requests.post(
                f"{self.base_url}/api/collections/_superusers/auth-with-password",
                json={"identity": email, "password": password},
                timeout=15,
            )
        except requests.exceptions.RequestException as err:
            print(f"\nLogin request failed: {err}")
            sys.exit(1)
        if not r.ok:
            print(f"\nLogin failed ({r.status_code}): {r.text}")
            sys.exit(1)
        self.token = r.json()["token"]
        print("Logged in.")

    def _headers(self):
        return {"Authorization": self.token} if self.token else {}

    def find_one(self, collection, filter_str):
        r = requests.get(
            f"{self.base_url}/api/collections/{collection}/records",
            params={"filter": filter_str, "perPage": 1},
            headers=self._headers(),
            timeout=15,
        )
        r.raise_for_status()
        items = r.json().get("items", [])
        return items[0] if items else None

    def create(self, collection, data):
        r = requests.post(
            f"{self.base_url}/api/collections/{collection}/records",
            json=data,
            headers=self._headers(),
            timeout=15,
        )
        if not r.ok:
            print(f"  ERROR creating {collection}: {r.status_code} {r.text}")
            print(f"  payload was: {data}")
            r.raise_for_status()
        return r.json()


def import_articles(wb, pb):
    ws = wb["Lagerbestand"]
    col = header_map(ws, LAGERBESTAND_HEADER_ROW)
    article_id_by_number = {}
    n_articles = 0
    n_items = 0

    for r in range(LAGERBESTAND_HEADER_ROW + 1, ws.max_row + 1):
        article_number = to_id_str(ws.cell(row=r, column=col["Artikel-Nr."]).value)
        if not article_number:
            continue

        existing = pb.find_one("articles", f'article_number = "{article_number}"')
        if existing:
            # NOTE: still fall through to the size/stock loop below - an
            # earlier run may have crashed after creating the article but
            # before finishing its article_items rows, so we must not skip
            # those. Each article_items row is checked/created individually.
            print(f"  articles: {article_number} already exists, skipping create")
            rec = existing
        else:
            data = {
                "article_number": article_number,
                "product_id": ws.cell(row=r, column=col["Handle"]).value or "",
                "name": ws.cell(row=r, column=col["Handle"]).value or "",
                "price": ws.cell(row=r, column=col["Preis"]).value or 0,
                "cost_price": ws.cell(row=r, column=col["Preis\nThömus"]).value or 0,
                "main_category": ws.cell(row=r, column=col["Rubrik"]).value or "",
                "category": ws.cell(row=r, column=col["Kategorie"]).value or "",
                "color_name": ws.cell(row=r, column=col["Artikel"]).value or "",
                "color_code": ws.cell(row=r, column=col["Farbe"]).value or "",
                "image_urls": parse_image_urls(ws.cell(row=r, column=col["URL Artikelbild"]).value),
                "embed_3d": ws.cell(row=r, column=col["3D Bild"]).value or "",
                "just_stock": ws.cell(row=r, column=col["JustStock"]).value or "no",
                "return_category": ws.cell(row=r, column=col["Rückgabe"]).value or "no",
                "description": ws.cell(row=r, column=col["Beschreibung"]).value or "",
                "notes": ws.cell(row=r, column=col["Bemerkungen Lagerbestand"]).value or "",
                "sort_order": ws.cell(row=r, column=col["Pos."]).value or 0,
            }
            rec = pb.create("articles", data)
            n_articles += 1
            print(f"  articles: created {article_number} ({data['name']} / {data['color_name']})")

        article_id_by_number[article_number] = rec["id"]

        for size_name in SIZE_COLUMNS:
            if size_name not in col:
                continue
            val = ws.cell(row=r, column=col[size_name]).value
            if val is None or isinstance(val, str):
                continue  # 'x' or blank -> product doesn't come in this size
            existing_item = pb.find_one(
                "article_items", f'article = "{rec["id"]}" && size = "{size_name}"'
            )
            if existing_item:
                continue
            pb.create("article_items", {
                "article": rec["id"],
                "size": size_name,
                "stock": int(val),
            })
            n_items += 1

    print(f"articles imported: {n_articles}, article_items imported: {n_items}")
    return article_id_by_number


def import_orders(wb, pb, article_id_by_number):
    ws = wb["Bestellungen"]
    col = header_map(ws, 1)

    orders = {}
    for r in range(2, ws.max_row + 1):
        oid = to_id_str(ws.cell(row=r, column=col["Order ID"]).value)
        if not oid:
            continue
        orders.setdefault(oid, []).append(r)

    n_orders = 0
    n_items = 0
    n_logs = 0

    for oid, rows in orders.items():
        first = rows[0]
        name_full = str(ws.cell(row=first, column=col["Name"]).value or "").strip()
        parts = re.split(r"\s+", name_full, maxsplit=1)
        buyer_name = parts[0] if parts else ""
        buyer_lastname = parts[1] if len(parts) > 1 else ""

        timestamp = ws.cell(row=first, column=col["Timestamp"]).value
        placed_at = timestamp.isoformat() if hasattr(timestamp, "isoformat") else ""

        existing_order = pb.find_one("orders", f'order_number = "{oid}"')

        if existing_order:
            order_rec = existing_order
            create_items = False
            print(f"  orders: {oid} already exists, skipping create")
        else:
            order_data = {
                "order_number": oid,
                "buyer_name": buyer_name or "unknown",
                "buyer_lastname": buyer_lastname,
                "buyer_email": ws.cell(row=first, column=col["Email"]).value or "",
                "kidzbike": bool(ws.cell(row=first, column=col["KidzBike"]).value),
                "comments": ws.cell(row=first, column=col["Comments"]).value or "",
                "payment_provider": "legacy",
                "payment_status": "confirmed",
                "amount_paid": ws.cell(row=first, column=col["Summe Pro Kunde"]).value or 0,
                "currency": ws.cell(row=first, column=col["Währung"]).value or "CHF",
                "placed_at": placed_at,
                # tri-state text field: "" = unknown. Legacy data doesn't tell
                # us whether an order was ever marked ready/picked up, so
                # leave it explicitly unknown rather than defaulting to "no"
                # (which would wrongly imply we know it wasn't).
                "ready": "",
                "picked_up": "",
            }
            order_rec = pb.create("orders", order_data)
            n_orders += 1
            create_items = True
            print(f"  orders: created {oid} ({buyer_name} {buyer_lastname})")

        # Backfill support: an earlier run of this script (before "logs"
        # support was added) may have already created this order without a
        # purchase log entry. Detect that and add the missing log now,
        # without re-creating order_items (those already exist).
        need_log = True
        if not create_items:
            already_logged = pb.find_one(
                "logs", f'order = "{order_rec["id"]}" && kind = "purchase"'
            )
            need_log = not already_logged

        if not create_items and not need_log:
            continue  # this order is fully done already, nothing left to do

        note_parts = []
        for r in rows:
            article_number = to_id_str(ws.cell(row=r, column=col["Artikelnummer"]).value)
            article_id = article_id_by_number.get(article_number)
            if not article_id:
                print(f"    WARNING: order {oid} references unknown article {article_number}, skipping line")
                continue

            size = ws.cell(row=r, column=col["Grösse"]).value or ""
            qty = ws.cell(row=r, column=col["Anzahl"]).value or 0

            if create_items:
                item_data = {
                    "order": order_rec["id"],
                    "article": article_id,
                    "size": size,
                    "color": ws.cell(row=r, column=col["Farbe"]).value or "",
                    "quantity": qty,
                    "unit_price": ws.cell(row=r, column=col["Preis pro Artikel"]).value or 0,
                    "price_paid": ws.cell(row=r, column=col["Tatsächlich eingegangen"]).value or 0,
                    "is_return": bool(ws.cell(row=r, column=col["isReturn"]).value),
                    "return_category": "",
                }
                pb.create("order_items", item_data)
                n_items += 1

            if need_log:
                note_parts.append(f"{article_number} x{qty} ({size})")

        if need_log:
            # so legacy orders show up in the dashboard's log/history view
            # too, same as orders placed through the new system - just as a
            # single "purchase" entry (we don't know ready/picked_up history
            # for these, see the ready/picked_up "" = unknown convention).
            action = "creating" if create_items else "backfilling missing"
            print(f"    logs: {action} purchase log for {oid}")
            pb.create("logs", {
                "order": order_rec["id"],
                "kind": "purchase",
                "note": ", ".join(note_parts),
                "placed_at": placed_at,
            })
            n_logs += 1

    print(f"orders imported: {n_orders}, order_items imported: {n_items}, logs imported: {n_logs}")


def main():
    if len(sys.argv) != 2:
        print(f"Usage: python {sys.argv[0]} <path-to-xlsx>")
        sys.exit(1)

    wb = openpyxl.load_workbook(sys.argv[1], data_only=True)

    pb = PocketBase(POCKETBASE_URL)
    pb.login()

    print("\nImporting articles + stock...")
    article_id_by_number = import_articles(wb, pb)

    print("\nImporting historical orders...")
    import_orders(wb, pb, article_id_by_number)

    print("\nDone.")


if __name__ == "__main__":
    main()
